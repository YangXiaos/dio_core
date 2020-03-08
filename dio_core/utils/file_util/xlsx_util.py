# @Time         : 18-6-17 下午9:09
# @Author       : DioMryang
# @File         : xlsx_util.py
# @Description  :
import logging
from collections import OrderedDict
from typing import List, Dict

import openpyxl


def write_rows2_xlsx(file_path, data=None):
    """
    以 xlsx 形式写入
    :param file_path:
    :param data:
    :return:
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    for ind, row in enumerate(data):
        for indChild, ele in enumerate(row):
            ws.cell(row=ind + 1, column=indChild + 1).value = ele
    wb.save(file_path)


def get_rows_from_xlsx(file_name: str) -> List[str]:
    """获取"""
    excel = openpyxl.load_workbook(file_name=file_name)
    for row in excel.active.iter_rows():
        yield list(map(lambda field: field.internal_value, row))


def python2xlsx(rows: List[dict], file_path: str, fields: List):
    """python对象 格式 2 xlxs"""
    wb = openpyxl.Workbook()
    ws = wb.active

    for row in rows:
        try:
            ws.append([row.get(field, "") for field in fields])
        except Exception as e:
            print(row.__str__() + "write fail")
    wb.save(file_path)
    wb.close()


def get_dict_from_xlsx(file_name: str):
    """获取"""
    excel = openpyxl.load_workbook(file_name=file_name)
    fields = []
    for row in excel.active.iter_rows():
        if not fields:
            fields = list(map(lambda field: field.internal_value, row))
            continue
        item = list(map(lambda field: field.internal_value, row))
        if len(item) == len(fields):
            yield OrderedDict({k: v for k, v in zip(fields, item)})
        else:
            logging.error("error item {}".format(item))


if __name__ == '__main__':
    # write_rows2_xlsx("xx.xlsx", [[1,2,3], [1,2,3]])
    rowsCmt = list(get_dict_from_xlsx("/home/changshuai/Documents/WXWork/1688850552307702/Cache/File/2019-06/【华为】twitter评论-69763条-20190610.xlsx"))
    rows = list(get_dict_from_xlsx("/home/changshuai/PycharmProjects/dio_core/dio_core/utils/teg/华为_twitter_搜索[HUAWEI]_changshuai_20190610.xlsx"))

    mapping = {}
    for row in rows:
        twitter_id = row['top_parent_twitter_id']
        mapping[twitter_id] = row["keyword"]

    for row in rowsCmt:
        if row["top_parent_twitter_id"] in mapping:
            row["keyword"] = mapping[row["top_parent_twitter_id"]]
        else:
            print(row)
    print()

